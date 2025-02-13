import streamlit as st
import pandas as pd
import random
from openpyxl import load_workbook
from fpdf import FPDF
from io import BytesIO

# Set up the page title and layout
st.set_page_config(page_title="Muziek Bingo", layout="centered")

# Function to set the style
def set_style():
    with open('./assets/styles.css') as f:
        css = f.read()
    st.markdown(f'<style>{css}</style>', unsafe_allow_html=True)

# Set the style
set_style()

# Title of the app
st.markdown('<div class="title">Muziek Bingo</div>', unsafe_allow_html=True)

# File uploader box
st.markdown('<div class="box">', unsafe_allow_html=True)
uploaded_file = st.file_uploader("Upload je afspeellijst in excel formaat")
if uploaded_file is not None:
    try:
        dataframe = pd.read_excel(uploaded_file)
        st.write(dataframe)
    except Exception as e:
        st.error(f"Error: {str(e)}")
else:
    st.text('Je hebt nog geen bestand geüpload')
st.markdown('</div>', unsafe_allow_html=True)

if uploaded_file is not None:
    playlist = pd.DataFrame(dataframe)

    # Random seed input box
    st.markdown('<div class="box">', unsafe_allow_html=True)
    random_seed = st.number_input('Vul hier je favoriete nummer in', step=1)
    st.write('Jouw favoriete nummer is: ', random_seed)
    seed_num = random_seed
    st.markdown('</div>', unsafe_allow_html=True)

    # Function to generate bingo cards
    def kaart_generator2(playlist, seed_num):
        random.seed(seed_num)
        nums = list(range(1, 51))
        random.shuffle(nums)
        playlist['random'] = nums
        new_df = playlist.sort_values("random")
        data = list(zip(new_df.iloc[0:5]['title_and_artist'],
                        new_df.iloc[5:10]['title_and_artist'],
                        new_df.iloc[10:15]['title_and_artist'],
                        new_df.iloc[15:20]['title_and_artist'],
                        new_df.iloc[20:25]['title_and_artist']))
        cols = ['B', 'I', 'N', 'G', 'O']

        df2 = pd.DataFrame(data, columns=cols)
        df2.at[2, 'N'] = "BINGO"

        return df2

    # Card number input box
    st.markdown('<div class="box">', unsafe_allow_html=True)
    card_num = st.number_input('Vul hier in hoeveel bingokaarten je wil genereren', step=1)
    st.write('Aantal bingokaarten ', card_num)
    st.markdown('</div>', unsafe_allow_html=True)

    def bingo_kaarten_generator2(playlist, aantal_kaarten, seed_num):
        kaarten_list = []
        for i in range(aantal_kaarten):
            kaarten_list.append(kaart_generator2(playlist, seed_num + i))
        return kaarten_list

    if card_num > 0:
        kaarten_list = bingo_kaarten_generator2(playlist, card_num, seed_num)

        @st.cache_data
        def convert_df(df):
            return df.to_csv(index=False, header=False, sep=';').encode('UTF-16')

        combined_df = pd.concat(kaarten_list, keys=range(1, card_num + 1), names=['Card', 'Row'])
        csv = convert_df(combined_df)

        st.download_button(
            label="Download data as CSV",
            data=csv,
            file_name='bingo_kaarten.csv',
            mime='text/csv',
        )

        # Display the first generated bingo card
        st.table(kaart_generator2(playlist, seed_num))

        # Display the pretty DataFrame in Streamlit
        st.dataframe(kaart_generator2(playlist, seed_num).style.set_properties(**{'text-align': 'center'}))

        # Create a PDF class inheriting from FPDF
        class PDF(FPDF):
            def header(self):
                self.set_font('Arial', 'B', 16)
                self.cell(0, 10, 'Bingo Cards', ln=True, align='C')
                self.ln(10)

            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

            def add_table(self, df, card_number):
                self.set_font('Arial', 'B', 12)
                self.cell(0, 10, f'Card {card_number}', ln=True, align='C')
                self.ln(5)

                self.set_font('Arial', '', 12)
                col_width = self.epw / 5  # distribute content evenly
                row_height = self.font_size * 1.5

                # Add table header
                for col_name in df.columns:
                    self.multi_cell(col_width, row_height, col_name, border=1, align='C', ln=3, max_line_height=self.font_size)
                self.ln(row_height)

                # Add table rows
                for row in df.itertuples(index=False):
                    for cell in row:
                        self.multi_cell(col_width, row_height, str(cell), border=1, align='C', ln=3, max_line_height=self.font_size)
                    self.ln(row_height)

        def download_as_pdf(cards):
            pdf = PDF(orientation='L', unit='mm', format='A4')
            pdf.add_page()
            for i, card in enumerate(cards, start=1):
                pdf.add_table(card, i)
                pdf.add_page()

            pdf_output = BytesIO()
            pdf.output(pdf_output)
            pdf_output.seek(0)
            return pdf_output

        if st.button('Download as PDF'):
            with st.spinner('Generating PDF...'):
                pdf_file = download_as_pdf(kaarten_list)
                st.success('Download Completed!')
                st.download_button(
                    label='Click to Download',
                    data=pdf_file,
                    file_name='bingo_cards.pdf',
                    mime='application/pdf'
                )

# import streamlit as st
# import pandas as pd
# import random
# from fpdf import FPDF
# from io import BytesIO

# # Set up the page title and layout
# st.set_page_config(page_title="Muziek Bingo", layout="centered")

# # Title of the app
# st.title("Muziek Bingo")

# # File uploader box
# uploaded_file = st.file_uploader("Upload je afspeellijst in Excel-formaat")
# if uploaded_file is not None:
#     try:
#         playlist = pd.read_excel(uploaded_file)
#         st.write(playlist)
#     except Exception as e:
#         st.error(f"Error: {str(e)}")
#         st.stop()
# else:
#     st.warning("Je hebt nog geen bestand geüpload")
#     st.stop()

# # Random seed input
# random_seed = st.number_input('Vul hier je favoriete nummer in', step=1, value=0)
# st.write('Jouw favoriete nummer is: ', random_seed)

# # Function to generate a single bingo card
# def kaart_generator(playlist, seed_num):
#     random.seed(seed_num)
#     shuffled_playlist = playlist.sample(frac=1, random_state=seed_num).reset_index(drop=True)
#     data = [shuffled_playlist.iloc[i * 5:(i + 1) * 5]['title_and_artist'].tolist() for i in range(6)]
    
#     df = pd.DataFrame(data, columns=['B', 'I', 'N', 'G', 'O'])
#     df.at[2, 'N'] = "BINGO"  # Center free space
#     return df

# # Input number of bingo cards
# card_num = st.number_input('Hoeveel bingokaarten wil je genereren?', step=1, min_value=1, value=1)
# st.write(f'Aantal bingokaarten: {card_num}')

# # Generate bingo cards
# kaarten_list = [kaart_generator(playlist, random_seed + i) for i in range(card_num)]

# # Display first generated bingo card
# st.table(kaarten_list[0])

# # PDF class with proper formatting
# class PDF(FPDF):
#     def header(self):
#         self.set_font('Arial', 'B', 16)
#         self.cell(0, 10, 'Bingo Cards', ln=True, align='C')
#         self.ln(5)

#     def footer(self):
#         self.set_y(-15)
#         self.set_font('Arial', 'I', 8)
#         self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

#     def add_table(self, df, card_number):
#         self.add_page()
#         self.set_font('Arial', 'B', 14)
#         self.cell(0, 10, f'Bingo Card {card_number}', ln=True, align='C')
#         self.ln(10)
        
#         table_size = 180  # Square table fitting within A4 landscape
#         col_width = table_size / 5
#         row_height = table_size / 6  # 6 rows for correct sizing
#         x_start = (self.w - table_size) / 2
        
#         # Table header
#         self.set_x(x_start)
#         for col_name in df.columns:
#             self.cell(col_width, row_height, col_name, border=1, align='C')
#         self.ln(row_height)
        
#         # Table rows
#         self.set_font('Arial', '', 10)
#         for row in df.itertuples(index=False):
#             self.set_x(x_start)
#             for cell in row:
#                 self.cell(col_width, row_height, str(cell), border=1, align='C')
#             self.ln(row_height)

# # Generate PDF function
# def generate_pdf(cards):
#     pdf = PDF(orientation='P', unit='mm', format='A4')  # Portrait mode for better layout
#     for i, card in enumerate(cards, start=1):
#         pdf.add_table(card, i)
#     pdf_output = BytesIO()
#     pdf.output(pdf_output)
#     pdf_output.seek(0)
#     return pdf_output

# # Button to download PDF
# if st.button('Download als PDF'):
#     with st.spinner('PDF wordt gegenereerd...'):
#         pdf_file = generate_pdf(kaarten_list)
#         st.download_button(
#             label='Klik hier om te downloaden',
#             data=pdf_file,
#             file_name='bingo_kaarten.pdf',
#             mime='application/pdf'
#         )
